import pandas as pd
import torch
from torch.utils.data import TensorDataset, DataLoader, RandomSampler, SequentialSampler
from transformers import BertModel,BertPreTrainedModel,BertTokenizer,AdamW, get_linear_schedule_with_warmup
import numpy as np
import time
import os
import torch.nn as nn
global EPOCHS, BATCH_SIZE_RATIO, SEQUENCE_LEN, LEARNING_RATE, TOKENIZER, MODEL_NAME
import torch.nn.functional as F
from openpyxl import load_workbook
projectnum = 2


global DATA_PATH
RESULT_FILE_PATH = r'Within_result.xlsx'
DATA_PATH = r'EXP_DATA/'
ROW_MAE, ROW_MMRE, ROW_PRED = 38, 39, 40

# TRAIN_TEST_FILE_PAIRS=os.listdir(DATA_PATH)
# TRAIN_TEST_FILE_PAIRS.sort(key = str.lower)
TRAIN_TEST_FILE_PAIRS = []
with open('exp_team_list.txt', 'r') as f:
    TRAIN_TEST_FILE_PAIRS = eval(f.read())

wb = load_workbook(RESULT_FILE_PATH)
sheet = wb.active
EPOCHS = 20
BATCH_SIZE_RATIO = 0.3
SEQUENCE_LEN = 20
LEARNING_RATE = 5e-4

# define device
global DEVICE
DEVICE = torch.device("cuda") if torch.cuda.is_available() else torch.device("cpu")


OUTPUT = ''
MODEL = None
DYNAMIC_BATCH = True
BATCH_SIZE = None
WITHIN_PROJECT = None
MAE_RECORDS = []
MDAE_RECORDS = []

def data_processing(file_pair):
    global BATCH_SIZE, BATCH_SIZE_RATIO, DATA_PATH, WITHIN_PROJECT, DYNAMIC_BATCH

    train_data = pd.DataFrame()
    fname = DATA_PATH + file_pair
    df = prepare_dataframe(fname)
    train_data = train_data.append(df)
        
    # data split
    if WITHIN_PROJECT:
        train_ex,train_labels,val_ex, val_labels, test_ex, test_labels = within_project_split(train_data)
    # define batch size dynamicalloutputsy based on training length
    if DYNAMIC_BATCH:
        BATCH_SIZE = int(len(train_ex) * BATCH_SIZE_RATIO)
    # tokenization

    train_ex=np.array(train_ex)
    train_ex=torch.tensor(train_ex)
    train_y = torch.tensor(train_labels.tolist()).type(torch.FloatTensor)
    train_dataloader = prepare_dataloader(train_ex, train_y, sampler_type='random')

    val_ex=np.array(val_ex)
    val_ex=torch.tensor(val_ex)
    val_y = torch.tensor(val_labels.tolist()).type(torch.FloatTensor)
    val_dataloader = prepare_dataloader(val_ex, val_y, sampler_type='sequential')
    
    # prepare testing datasets
    all_test_dataloader = []
    test_file_names = []
    if WITHIN_PROJECT:
        test_ex=np.array(test_ex)
        test_ex=torch.tensor(test_ex)
        test_y = torch.tensor(test_labels.tolist()).type(torch.FloatTensor)
        test_dataloader = prepare_dataloader(test_ex, test_y, sampler_type='sequential')
        all_test_dataloader.append(test_dataloader)
        test_file_names.append(file_pair)
        return file_pair, train_dataloader, val_dataloader, all_test_dataloader, test_file_names



def prepare_dataframe(file_name):
    data = pd.read_csv(file_name)
    order=[ 'contributors','creater_ar_count','developer_ar_count','tester_ar_count','committer_qualification_level',
                                    'detail_version_name','author_commit_num','author_modify_file_num','author_create_mr_num','author_update_mr_num','author_mr_commit_num', 'dev_dts_num','test_dts_num','actual_effort']
    data=data[order]
    data = data.fillna(0)
    return pd.DataFrame(data=data)


def prepare_dataloader(seq, y, sampler_type):
    global BATCH_SIZE
    tensor_dataset = TensorDataset(seq, y)
    if sampler_type == 'random':
        sampler = RandomSampler(tensor_dataset)
    elif sampler_type == 'sequential':
        sampler = SequentialSampler(tensor_dataset)
    dataloader = DataLoader(tensor_dataset, sampler=sampler, batch_size=BATCH_SIZE)
    return dataloader


def within_project_split(data):
    print('within project split!')
    train_val_split_point = int(len(data) * 0.6)
    val_test_split_point = int(len(data) * 0.8)
    train_ex=data.iloc[:train_val_split_point,0:13]
    train_labels = (data['actual_effort'][:train_val_split_point])
    val_ex=data.iloc[train_val_split_point:val_test_split_point,0:13]
    val_labels = (data['actual_effort'][train_val_split_point:val_test_split_point])
    test_ex=data.iloc[val_test_split_point:,0:13]
    test_labels = (data['actual_effort'][val_test_split_point:])
    return train_ex,train_labels,val_ex, val_labels, test_ex, test_labels


class MYNET(nn.Module):
    def __init__(self):
        super(MYNET, self).__init__()
        self.dropout = nn.Dropout(0.3)
        self.score = nn.Linear(13, 1)

    def forward(self, inputs,labels=None):
        inputs = inputs.float()
        inputs = self.dropout(inputs)
        logit = self.score(inputs)
        return logit

def train_eval_test(file_pair, train_dataloader, val_dataloader, all_test_dataloader, model, test_file_names):
    global LEARNING_RATE, EPOCHS, MAE_RECORDS, MDAE_RECORDS, DEVICE
    optimizer = AdamW(MODEL.parameters(), lr=LEARNING_RATE)    
    # Total number of training steps is [number of batches] x [number of epochs]
    total_steps = len(train_dataloader) * EPOCHS
    # Create the learning rate scheduler
    scheduler = get_linear_schedule_with_warmup(optimizer, num_warmup_steps=0, num_training_steps=total_steps)
    print("Start training for ", file_pair, ".....")
    training_start_time = time.time()
    

    min_eval_loss_epoch = [10000, 0]
    
    time_records = []
    MAE_RECORDS = []
    MSE_RECORDS = []
    MDAE_RECORDS = []
    MMRE_RECORDS = []
    PRED_RECPRDS=[]
    start_time = time.time()
    loss_fct = nn.L1Loss()
    for e in range(EPOCHS):
        # ---TRAINING---
        # clean GPU memory
        torch.cuda.empty_cache()
        print(">>> epoch ", e)
        # set model into train mode
        model.train()
        total_train_loss = 0
        for step, batch in enumerate(train_dataloader):            
            b_input_ids = batch[0].to(DEVICE)
            b_labels = batch[1].to(DEVICE)
            model.zero_grad()
            result = model(b_input_ids, 
                           labels=b_labels,
                           )
            loss = loss_fct(result,b_labels)
            logits=result
            total_train_loss += loss.item()  
            loss.backward() 
            torch.nn.utils.clip_grad_norm_(model.parameters(), 1.0)
            optimizer.step()
            scheduler.step()
            # clean memory
            del step, batch, b_input_ids, b_labels, result, loss, logits

        avg_train_loss = total_train_loss / len(train_dataloader)
        print(" Average training MAE loss: {0:.2f}".format(avg_train_loss))
        # clean memory
        del avg_train_loss, total_train_loss
        
        time_records.append(time.time() - start_time)
        
        # ---EVAL---
        print("-")
        # set model into eval mode
        model.eval()
        total_eval_loss = 0
        for batch in val_dataloader:            
            b_input_ids = batch[0].to(DEVICE)
            b_labels = batch[1].to(DEVICE)
            model.zero_grad()
            result = model(b_input_ids, 
                           labels=b_labels,
                           )
            loss = loss_fct(result,b_labels)
            logits = result
            total_eval_loss += loss.item()  
            # clean memory
            del b_input_ids, b_labels, batch, result, loss, logits
        avg_eval_loss = total_eval_loss / len(val_dataloader)
        print(" Average eval MAE loss: {0:.2f}".format(avg_eval_loss))
        
        if avg_eval_loss <= min_eval_loss_epoch[0]:
            min_eval_loss_epoch[0] = avg_eval_loss
            min_eval_loss_epoch[1] = e
        
        # clean memory
        del avg_eval_loss, total_eval_loss
        # save model state to dict
        
        print("===============================")
        
        # testing on holdout data
        index = 0
        for test_dataloader in all_test_dataloader:
            test_file_name = test_file_names[index]
            index += 1
            testing_start_time = time.time()
            predictions = []
            true_labels = []
            for batch in test_dataloader:
                batch = tuple(t.to(DEVICE) for t in batch)
                b_input_ids, b_labels = batch
                with torch.no_grad():
                    logits = model(b_input_ids)
                logits = logits.detach().cpu().numpy()
                label_ids = b_labels.to('cpu').numpy()
                predictions.append(logits)
                true_labels.append(label_ids)
            # calculate errors
            total_distance = 0
            total_se = 0
            total_mre = 0
            m=0
            distance_records = []
            total_data_point=0
            for i in range(len(predictions)):
                total_data_point+=len(predictions[i])
            for i in range(len(predictions)):
                for j in range(len(predictions[i])):
                    distance = abs(predictions[i][j] - true_labels[i][j])
                    print("预测值："+str(predictions[i][j])+'    实际值:'+str(true_labels[i][j]))
                    se=(predictions[i][j] - true_labels[i][j])**2
                    if(true_labels[i][j]>0):
                        mre=abs(predictions[i][j] - true_labels[i][j])/true_labels[i][j]
                    else:
                        mre=(abs(predictions[i][j] - true_labels[i][j])+1)/(true_labels[i][j]+1)
                    if mre<0.5:
                        m+=1
                    total_se+=se
                    total_mre+=mre
                    total_distance += distance
                    distance_records.append(distance)
            MAE = total_distance / total_data_point
            MSE = (total_se**(0.5))/total_data_point
            MMRE= total_mre / total_data_point
            MdAE = np.median(np.array(distance_records)) 
            PRED=m/total_data_point
            MAE_RECORDS.append(MAE)
            MSE_RECORDS.append(MSE)
            MDAE_RECORDS.append(MdAE)
            MMRE_RECORDS.append(MMRE)
            PRED_RECPRDS.append(PRED)
            
            global OUTPUT
            OUTPUT +=  'Epochs ' + str(e) + '\n'
            OUTPUT += 'MAE: ' + str(MAE) + '\n'
            OUTPUT += 'MdAE: ' + str(MdAE) + '\n'
            OUTPUT += 'MSE: ' + str(MSE) + '\n'
            OUTPUT += 'MMRE: ' + str(MMRE) + '\n'
            OUTPUT += 'PRED: ' + str(PRED) + '\n\n'
            print('MAE: ', MAE)
            print('MdAE: ', MdAE)
            print('MSE: ', MSE)
            print('MMRE: ', MMRE)
            print('PRED: ', PRED)
    
            
    OUTPUT +=str(MAE_RECORDS[min_eval_loss_epoch[1]]) + '\n'+str(MSE_RECORDS[min_eval_loss_epoch[1]]) + '\n'+str(MMRE_RECORDS[min_eval_loss_epoch[1]]) + '\n'+ str(PRED_RECPRDS[min_eval_loss_epoch[1]]) + '\n'
    OUTPUT += 'training time: ' + str(time_records[min_eval_loss_epoch[1]]) + '\n'
    OUTPUT += 'Epochs: ' + str(min_eval_loss_epoch[1]) +'\n'
    global BATCH_SIZE
    OUTPUT += 'batch size: ' + str(BATCH_SIZE)
    print('all done for one project')
    sheet.cell(row=ROW_MAE, column=projectnum).value = MAE_RECORDS[min_eval_loss_epoch[1]][0]
    # sheet.cell(row=28, column=projectnum).value = MSE_RECORDS[min_eval_loss_epoch[1]][0]
    sheet.cell(row=ROW_MMRE, column=projectnum).value = MMRE_RECORDS[min_eval_loss_epoch[1]][0]
    sheet.cell(row=ROW_PRED, column=projectnum).value = PRED_RECPRDS[min_eval_loss_epoch[1]]
    wb.save(RESULT_FILE_PATH)

    
WITHIN_PROJECT = True
def main():
    global TRAIN_TEST_FILE_PAIRS, MODEL, TOKENIZER, MODEL_NAME
    
    sheet.cell(row=ROW_MAE-1, column=1).value = '纯专家特征'
    sheet.cell(row=ROW_MAE, column=1).value = 'MAE'
    sheet.cell(row=ROW_MMRE, column=1).value = 'MMRE'
    sheet.cell(row=ROW_PRED, column=1).value = 'PRED'
    
    for file in TRAIN_TEST_FILE_PAIRS:
        MODEL = MYNET()
        MODEL=MODEL.cuda()
        file_pair, train_dataloader, val_dataloader, all_test_dataloader, test_file_names = data_processing(file_pair=file)
        train_eval_test(file_pair, train_dataloader, val_dataloader, all_test_dataloader, MODEL, test_file_names)
        del MODEL
        torch.cuda.empty_cache()            
        global OUTPUT
        with open('./middle_output/' + str(file_pair[:-4]) +'.txt', 'w+') as f:
            f.writelines(OUTPUT)
            print('results have been written into a text file!')
            OUTPUT = ""
        global projectnum
        projectnum=projectnum+1

                
if __name__ == "__main__":
    main()